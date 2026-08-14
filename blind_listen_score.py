#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
盲听音频评分工具（单文件版）

依赖安装:
    pip install pygame openpyxl

用法:
    1. 开发：将各场景 wav 放入脚本同级 music/A、music/B ... 目录
       python blind_listen_score.py
    2. 发布：运行 build_exe.bat，生成单文件 exe（music 内嵌，用户不可直接替换）

跨平台开发: Windows / macOS / Linux（tkinter + pygame）
打包目标: Windows exe（PyInstaller --onefile）
"""

from __future__ import annotations

import array
import os
import struct
import sys
import tempfile
import threading
import wave
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# WAV 格式常量（Python wave 模块不支持 format=3 的 IEEE float）
WAVE_FORMAT_PCM = 1
WAVE_FORMAT_IEEE_FLOAT = 3
WAVE_FORMAT_EXTENSIBLE = 0xFFFE

# ---------------------------------------------------------------------------
# 依赖检查
# ---------------------------------------------------------------------------
try:
    import pygame
except ImportError:
    print("请先安装 pygame:  pip install pygame")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("请先安装 openpyxl:  pip install openpyxl")
    sys.exit(1)


# =============================================================================
# 配置宏定义（可按需修改）
# =============================================================================

# 音频根目录名（相对本脚本所在目录）
MUSIC_DIR_NAME = "music"

# 八个场景文件夹名（顺序即测评顺序）
SCENE_FOLDERS = ["A", "B", "C", "D", "E", "F", "G", "H"]

# 场景对外展示名（与 SCENE_FOLDERS 一一对应；可改成中文场景名）
SCENE_DISPLAY_NAMES = [
    "场景1:人声质量评分",
    "场景2:噪声场景-1",
    "场景3:噪声场景-2",
    "场景4:噪声场景-3",
    "场景5:噪声场景-4",
    "场景6:噪声场景-5",
    "场景7:噪声场景-6",
    "场景8:噪声场景-7",
]

# 第几个场景为人声质量评分（0-based）；该场景仅评 1 个维度
VOICE_QUALITY_SCENE_INDEX = 0

# 每个场景期望的音频数量
NUM_AUDIO_PER_SCENE = 10

# 支持的音频扩展名（盲听仅用 wav）
AUDIO_EXTENSIONS = (".wav", ".WAV")

# 界面编号位数，如 01、02
CODE_LABEL_WIDTH = 2

# 场景1：人声质量（仅 1 维）
VOICE_SCENE_DIMENSIONS: List[Tuple[str, str]] = [
    ("speech_quality", "语音质量（清晰度&还原度）"),
]

# 场景2~8：噪声场景（3 维）
NOISE_SCENE_DIMENSIONS: List[Tuple[str, str]] = [
    ("s_mos", "S_mos:人声信号（只看语音失真）"),
    ("n_mos", "N_MOS:背景噪声（只看噪声扰人程度）"),
    ("g_mos", "G_MOS:总体质量（综合听感）"),
]

# Excel 统一导出的全部评分列（本场景不适用的维度留空）
ALL_SCORE_DIMENSIONS: List[Tuple[str, str]] = (
    VOICE_SCENE_DIMENSIONS + NOISE_SCENE_DIMENSIONS
)


def dimensions_for_scene(scene_index: int) -> List[Tuple[str, str]]:
    """按场景返回需要填写的评分维度。"""
    if scene_index == VOICE_QUALITY_SCENE_INDEX:
        return VOICE_SCENE_DIMENSIONS
    return NOISE_SCENE_DIMENSIONS


def scene_type_label(scene_index: int) -> str:
    return "人声质量" if scene_index == VOICE_QUALITY_SCENE_INDEX else "噪声场景"


# ---------------------------------------------------------------------------
# 评分标准表（打分界面右侧常驻展示）
# 只需改 headers / rows 文案即可；rows 每一行对应表格一行
# ---------------------------------------------------------------------------

# 场景1：人声质量（示例已填 5/4 分，其余请自行补全）
VOICE_SCORE_RUBRIC: Dict[str, object] = {
    "title": "人声质量评分标准",
    "headers": ["分数", "清晰度", "还原度"],
    "rows": [
        [
            "5 优秀",
            "吐字字字清晰，无模糊吞字，无噪声失真，听音毫不费力",
            "音色高度还原原声，气息细节完整，听感自然真实，无畸变",
        ],
        [
            "4 良好",
            "大部分字词清晰，仅个别轻微模糊，微弱噪声失真，不影响理解",
            "大部分字词清晰，仅个别轻微模糊，微弱噪声失真，不影响理解",
        ],
        ["3 中等", "（待填写）", "（待填写）"],
        ["2 较差", "（待填写）", "（待填写）"],
        ["1 很差", "（待填写）", "（待填写）"],
        # ["0 分", "（待填写）", "（待填写）"],
    ],
}

# 场景2~8：噪声场景（三列维度，请自行填写各分行描述）
NOISE_SCORE_RUBRIC: Dict[str, object] = {
    "title": "噪声场景评分标准",
    "headers": ["分数", "S_MOS", "N_MOS", "G_MOS"],
    "rows": [
        ["5 优秀", "（待填写）", "（待填写）", "（待填写）"],
        ["4 良好", "（待填写）", "（待填写）", "（待填写）"],
        ["3 中等", "（待填写）", "（待填写）", "（待填写）"],
        ["2 较差", "（待填写）", "（待填写）", "（待填写）"],
        ["1 很差", "（待填写）", "（待填写）", "（待填写）"],
        # ["0 分", "（待填写）", "（待填写）", "（待填写）"],
    ],
}


def rubric_for_scene(scene_index: int) -> Dict[str, object]:
    if scene_index == VOICE_QUALITY_SCENE_INDEX:
        return VOICE_SCORE_RUBRIC
    return NOISE_SCORE_RUBRIC


# 分数范围与精度
SCORE_MIN = 0.0
SCORE_MAX = 5.0
SCORE_STEP = 0.1

# Excel 输出文件名前缀
EXCEL_FILENAME_PREFIX = "盲听评分结果"

# 进度条刷新间隔（毫秒）
PROGRESS_TICK_MS = 100

# ---------------------------------------------------------------------------
# 调试宏开关
# 1 = 只认真填写第 1 个场景；点击完成后，将其余场景按同编号(01~10)
#     复制评分并直接导出 Excel（噪声场景维度不同时，用第1场景分数填入各维便于测表）
# 0 = 正式测评（需逐场景填写全部评分）
# ---------------------------------------------------------------------------
DEBUG_COPY_FIRST_SCENE_SCORES = 0


# =============================================================================
# 工具函数
# =============================================================================

def resource_dir() -> Path:
    """
    只读资源目录（内置 music 等）。
    - 开发运行: 脚本所在目录
    - PyInstaller onefile/onedir: sys._MEIPASS（打包进 exe 的解压/资源目录）
    """
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            return Path(meipass)
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def writable_dir() -> Path:
    """
    可写目录（导出 Excel 等）。
    - 开发运行: 脚本所在目录
    - 打包后: exe 所在目录（不要写到 _MEIPASS）
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def app_base_dir() -> Path:
    """兼容旧调用：默认指向可写目录。读音频请用 music_root()。"""
    return writable_dir()


def music_root() -> Path:
    """音频根目录：打包后从内嵌资源读取，用户无法通过替换 exe 旁文件夹篡改。"""
    return resource_dir() / MUSIC_DIR_NAME


def audio_relpath(path: Path) -> str:
    """Excel 中记录的相对路径（相对资源根；失败则仅文件名）。"""
    try:
        return str(path.relative_to(resource_dir()))
    except ValueError:
        return path.name


def list_scene_wavs(scene_folder: str) -> List[Path]:
    """按文件名排序列出场景下的 wav；数量由目录实际内容决定。"""
    folder = music_root() / scene_folder
    if not folder.is_dir():
        return []
    files = [
        p for p in folder.iterdir()
        if p.is_file() and p.suffix in AUDIO_EXTENSIONS
    ]
    return sorted(files, key=lambda p: p.name.lower())


def parse_wav_info(path: Path) -> dict:
    """
    解析 WAV 头信息，支持:
      - format 1: PCM 整型
      - format 3: IEEE float（你遇到的 unknown format: 3）
      - format 0xFFFE: WAVE_FORMAT_EXTENSIBLE
    """
    with open(path, "rb") as f:
        header = f.read(12)
        if len(header) < 12 or header[0:4] != b"RIFF" or header[8:12] != b"WAVE":
            raise ValueError("不是有效的 WAV / RIFF 文件")

        audio_format = None
        channels = None
        sample_rate = None
        byte_rate = None
        bits_per_sample = None
        data_size = None
        data_offset = None

        while True:
            chunk_hdr = f.read(8)
            if len(chunk_hdr) < 8:
                break
            chunk_id, chunk_size = struct.unpack("<4sI", chunk_hdr)
            # 防止异常超大 chunk
            if chunk_size > 1 << 31:
                raise ValueError(f"异常 WAV chunk 大小: {chunk_id!r}")

            if chunk_id == b"fmt ":
                fmt = f.read(chunk_size)
                if len(fmt) < 16:
                    raise ValueError("fmt chunk 过短")
                audio_format, channels, sample_rate, byte_rate, _block_align, bits_per_sample = (
                    struct.unpack("<HHIIHH", fmt[:16])
                )
                # Extensible: 实际格式在 SubFormat GUID 前 2 字节
                if audio_format == WAVE_FORMAT_EXTENSIBLE and len(fmt) >= 26:
                    audio_format = struct.unpack("<H", fmt[24:26])[0]
            elif chunk_id == b"data":
                data_offset = f.tell()
                data_size = chunk_size
                f.seek(chunk_size, os.SEEK_CUR)
            else:
                f.seek(chunk_size, os.SEEK_CUR)

            # chunk 偶数字节对齐
            if chunk_size % 2 == 1:
                f.seek(1, os.SEEK_CUR)

    if None in (audio_format, channels, sample_rate, bits_per_sample, data_size, data_offset):
        raise ValueError("WAV 缺少 fmt/data 信息")
    if sample_rate <= 0 or channels <= 0 or bits_per_sample <= 0:
        raise ValueError("WAV 参数非法")

    bytes_per_sample = bits_per_sample // 8
    if bytes_per_sample <= 0:
        raise ValueError(f"不支持的位深: {bits_per_sample}")
    frame_size = channels * bytes_per_sample
    nframes = data_size // frame_size
    duration = nframes / float(sample_rate)

    return {
        "audio_format": audio_format,
        "channels": channels,
        "sample_rate": sample_rate,
        "byte_rate": byte_rate,
        "bits_per_sample": bits_per_sample,
        "data_size": data_size,
        "data_offset": data_offset,
        "nframes": nframes,
        "duration": duration,
    }


def wav_duration_seconds(path: Path) -> float:
    """读取 wav 时长（秒），兼容 PCM / IEEE float。"""
    return float(parse_wav_info(path)["duration"])


def _read_wav_pcm16_bytes(path: Path) -> Tuple[bytes, int, int]:
    """
    读取任意常见 WAV，返回 (int16 PCM 交错字节流, sample_rate, channels)。
    float32 / float64 会裁剪到 [-1, 1] 再转 int16。
    """
    info = parse_wav_info(path)
    fmt = info["audio_format"]
    channels = info["channels"]
    rate = info["sample_rate"]
    bps = info["bits_per_sample"]

    with open(path, "rb") as f:
        f.seek(info["data_offset"])
        raw = f.read(info["data_size"])

    if fmt == WAVE_FORMAT_PCM and bps == 16:
        return raw, rate, channels

    if fmt == WAVE_FORMAT_PCM and bps == 8:
        # 8bit PCM 为无符号
        samples = array.array("B", raw)
        out = array.array("h", ((s - 128) << 8 for s in samples))
        return out.tobytes(), rate, channels

    if fmt == WAVE_FORMAT_PCM and bps == 24:
        n = len(raw) // 3
        out = array.array("h")
        for i in range(n):
            b0, b1, b2 = raw[i * 3 : i * 3 + 3]
            val = b0 | (b1 << 8) | (b2 << 16)
            if val & 0x800000:
                val -= 0x1000000
            out.append(max(-32768, min(32767, val >> 8)))
        return out.tobytes(), rate, channels

    if fmt == WAVE_FORMAT_PCM and bps == 32:
        samples = array.array("i")
        samples.frombytes(raw[: len(raw) - (len(raw) % 4)])
        if sys.byteorder != "little":
            samples.byteswap()
        out = array.array("h", (max(-32768, min(32767, s >> 16)) for s in samples))
        return out.tobytes(), rate, channels

    if fmt == WAVE_FORMAT_IEEE_FLOAT and bps == 32:
        samples = array.array("f")
        samples.frombytes(raw[: len(raw) - (len(raw) % 4)])
        if sys.byteorder != "little":
            samples.byteswap()
        out = array.array(
            "h",
            (max(-32768, min(32767, int(max(-1.0, min(1.0, s)) * 32767.0))) for s in samples),
        )
        return out.tobytes(), rate, channels

    if fmt == WAVE_FORMAT_IEEE_FLOAT and bps == 64:
        samples = array.array("d")
        samples.frombytes(raw[: len(raw) - (len(raw) % 8)])
        if sys.byteorder != "little":
            samples.byteswap()
        out = array.array(
            "h",
            (max(-32768, min(32767, int(max(-1.0, min(1.0, s)) * 32767.0))) for s in samples),
        )
        return out.tobytes(), rate, channels

    raise ValueError(f"暂不支持的 WAV 格式: format={fmt}, bits={bps}")


def ensure_playable_wav(path: Path, cache_dir: Path) -> Path:
    """
    返回 pygame 可直接播放的 PCM16 wav 路径。
    对 IEEE float（format=3）等会转换并缓存到临时目录。
    """
    info = parse_wav_info(path)
    if info["audio_format"] == WAVE_FORMAT_PCM and info["bits_per_sample"] == 16:
        return path

    cache_dir.mkdir(parents=True, exist_ok=True)
    # 用源文件签名避免重复转换
    stamp = f"{path.stat().st_mtime_ns}_{path.stat().st_size}"
    out = cache_dir / f"{path.stem}__pcm16__{stamp}.wav"
    if out.is_file():
        return out

    pcm, rate, channels = _read_wav_pcm16_bytes(path)
    with wave.open(str(out), "wb") as wf:
        wf.setnchannels(channels)
        wf.setsampwidth(2)
        wf.setframerate(rate)
        wf.writeframes(pcm)
    return out


class AudioDeviceError(RuntimeError):
    """无可用音频输出设备，或初始化超时/失败。"""


def _mixer_init_once(driver: Optional[str]) -> None:
    if driver:
        os.environ["SDL_AUDIODRIVER"] = driver
    else:
        os.environ.pop("SDL_AUDIODRIVER", None)
    try:
        pygame.mixer.quit()
    except Exception:
        pass
    pygame.mixer.init(frequency=44100, size=-16, channels=2, buffer=2048)


def init_pygame_mixer(timeout_sec: float = 4.0) -> str:
    """
    初始化音频设备；Windows 上 WASAPI 失败时回退 directsound / winmm。
    单次尝试超过 timeout_sec 视为无设备/卡死，抛出 AudioDeviceError。
    返回实际使用的驱动名（可能为空字符串表示默认）。
    """
    drivers: List[Optional[str]]
    if sys.platform.startswith("win"):
        drivers = ["directsound", "winmm", "wasapi", None]
    elif sys.platform == "darwin":
        drivers = ["coreaudio", None]
    else:
        drivers = ["pulseaudio", "alsa", None]

    last_err: Optional[BaseException] = None
    for driver in drivers:
        box: Dict[str, object] = {"ok": False, "err": None}

        def _worker(d: Optional[str] = driver) -> None:
            try:
                _mixer_init_once(d)
                box["ok"] = True
            except BaseException as exc:  # noqa: BLE001 - 需捕获 pygame 底层异常
                box["err"] = exc

        th = threading.Thread(target=_worker, daemon=True)
        th.start()
        th.join(timeout_sec)
        if th.is_alive():
            # 初始化卡住通常就是没有输出设备，直接提示，避免多驱动轮询拖很久
            raise AudioDeviceError(
                "音频设备初始化超时，可能未连接耳机/扬声器。\n"
                "请连接音频设备后重新打开本程序。"
            )
        if box["ok"]:
            return driver or ""
        last_err = box["err"] if isinstance(box["err"], BaseException) else last_err

    raise AudioDeviceError(
        "未检测到可用的音频输出设备。\n"
        "请连接耳机或扬声器后，重新打开本程序。\n\n"
        f"详细信息: {last_err}"
    )


def show_audio_device_error(detail: object = None) -> None:
    """弹窗提示连接音频设备（不依赖主窗口）。"""
    msg = (
        "未检测到音频设备。\n\n"
        "请连接耳机或扬声器后，重新打开本程序。"
    )
    if detail:
        msg += f"\n\n详细信息:\n{detail}"
    root = tk.Tk()
    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except tk.TclError:
        pass
    messagebox.showerror("请连接音频设备", msg, parent=root)
    try:
        root.destroy()
    except tk.TclError:
        pass


def format_time(seconds: float) -> str:
    seconds = max(0.0, float(seconds))
    m = int(seconds) // 60
    s = int(seconds) % 60
    return f"{m:02d}:{s:02d}"


def quantize_score(value: float) -> float:
    """量化到 SCORE_STEP 精度。"""
    steps = round(value / SCORE_STEP)
    return round(steps * SCORE_STEP, 1)


def is_valid_score(text: str) -> bool:
    """校验：0~5，步进 0.1（如 0.1、3.5 合法；5.1、3.22 非法）。"""
    try:
        v = float(text.strip())
    except ValueError:
        return False
    if v < SCORE_MIN or v > SCORE_MAX:
        return False
    q = quantize_score(v)
    return abs(v - q) < 1e-9


def code_label(index: int) -> str:
    """0-based index -> '01' 形式。"""
    return f"{index + 1:0{CODE_LABEL_WIDTH}d}"


# =============================================================================
# 音频播放器（pygame，支持拖动定位）
# =============================================================================

class WavPlayer:
    """同一时刻只播放一条音频。"""

    def __init__(self) -> None:
        self._driver = init_pygame_mixer()
        self.path: Optional[Path] = None
        self._play_path: Optional[Path] = None
        self.duration: float = 0.0
        self._seek_base: float = 0.0  # play(start=...) 的起点
        self._paused: bool = False
        self._pause_pos: float = 0.0
        self._cache_dir = Path(tempfile.gettempdir()) / "blind_listen_pcm_cache"

    def load(self, path: Path, duration: float) -> None:
        self.stop()
        self.path = path
        self.duration = duration
        self._seek_base = 0.0
        self._paused = False
        self._pause_pos = 0.0
        # float WAV 等先转成 PCM16，pygame 才能稳定播放
        self._play_path = ensure_playable_wav(path, self._cache_dir)
        pygame.mixer.music.load(str(self._play_path))

    def play(self, start: float = 0.0) -> None:
        if self.path is None or self._play_path is None:
            return
        start = max(0.0, min(start, max(0.0, self.duration - 0.05)))
        self._seek_base = start
        self._paused = False
        # pygame 2+: play(start=seconds) 对 wav 可用
        pygame.mixer.music.play(start=start)

    def pause(self) -> None:
        if not pygame.mixer.music.get_busy() and not self._paused:
            return
        self._pause_pos = self.get_position()
        pygame.mixer.music.pause()
        self._paused = True

    def unpause(self) -> None:
        if self._paused:
            # 部分平台 pause/unpause 对 set_pos 不稳定，用重新 play 更稳妥
            self.play(self._pause_pos)
            self._paused = False

    def stop(self) -> None:
        try:
            pygame.mixer.music.stop()
        except pygame.error:
            pass
        self._seek_base = 0.0
        self._paused = False
        self._pause_pos = 0.0

    def is_playing(self) -> bool:
        return (not self._paused) and pygame.mixer.music.get_busy()

    def is_paused(self) -> bool:
        return self._paused

    def get_position(self) -> float:
        if self._paused:
            return self._pause_pos
        if not pygame.mixer.music.get_busy():
            # 播放结束
            if self.path is not None and self._seek_base > 0:
                # 刚结束时 get_pos 可能为 -1
                return self.duration
            pos_ms = pygame.mixer.music.get_pos()
            if pos_ms < 0:
                return self.duration if self._seek_base > 0 or self.path else 0.0
            return min(self.duration, self._seek_base + pos_ms / 1000.0)
        pos_ms = pygame.mixer.music.get_pos()
        if pos_ms < 0:
            return self._seek_base
        return min(self.duration, self._seek_base + pos_ms / 1000.0)

    def seek(self, seconds: float) -> None:
        seconds = max(0.0, min(float(seconds), max(0.0, self.duration - 0.05) if self.duration > 0 else 0.0))
        if self.is_paused():
            # 暂停时只更新位置，不恢复播放
            self._pause_pos = seconds
            self._seek_base = seconds
            return
        if self.is_playing():
            self.play(seconds)
            return
        # 未在播：仅记录起点，供下次 play 使用
        self._seek_base = seconds
        self._pause_pos = seconds

    def shutdown(self) -> None:
        self.stop()
        try:
            pygame.mixer.quit()
        except Exception:
            pass

# =============================================================================
# 单条音频行 UI
# =============================================================================

class AudioRow:
    def __init__(
        self,
        parent: tk.Widget,
        index: int,
        path: Path,
        player: WavPlayer,
        on_play_request,
        dimensions: Optional[List[Tuple[str, str]]] = None,
    ) -> None:
        self.index = index
        self.path = path
        self.player = player
        self.on_play_request = on_play_request
        self.dimensions = list(dimensions or NOISE_SCENE_DIMENSIONS)
        self.code = code_label(index)
        self.duration = 0.0
        self._seeking = False
        self._active = False  # 当前是否由本行占用播放器

        try:
            self.duration = wav_duration_seconds(path)
        except Exception as exc:
            messagebox.showwarning(
                "音频读取失败",
                f"{path.name}\n{exc}\n\n"
                "提示: 若曾出现 unknown format: 3，说明是 IEEE float WAV，"
                "当前版本已支持该格式，请确认已保存最新脚本后重试。",
            )
            self.duration = 0.0

        self.frame = ttk.LabelFrame(parent, text=f"音频 {self.code}", padding=6)

        top = ttk.Frame(self.frame)
        top.pack(fill=tk.X)

        self.btn_play = ttk.Button(top, text="▶ 播放", width=10, command=self.toggle_play)
        self.btn_play.pack(side=tk.LEFT, padx=(0, 6))

        self.progress = ttk.Scale(
            top,
            from_=0.0,
            to=max(self.duration, 0.1),
            orient=tk.HORIZONTAL,
            command=self._on_seek_drag,
        )
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        # ttk.Scale 多数主题不支持点槽跳转；用鼠标坐标换算实现点击/拖槽定位
        self.progress.bind("<ButtonPress-1>", self._seek_press)
        self.progress.bind("<B1-Motion>", self._seek_motion)
        self.progress.bind("<ButtonRelease-1>", self._seek_release)

        self.time_var = tk.StringVar(value=f"00:00 / {format_time(self.duration)}")
        ttk.Label(top, textvariable=self.time_var, width=14).pack(side=tk.LEFT)

        score_frame = ttk.Frame(self.frame)
        score_frame.pack(fill=tk.X, pady=(8, 0))

        self.score_vars: Dict[str, tk.StringVar] = {}
        self.score_widgets: Dict[str, ttk.Entry] = {}

        vcmd = (self.frame.register(self._validate_score_key), "%P")

        for col, (key, title) in enumerate(self.dimensions):
            cell = ttk.Frame(score_frame)
            cell.grid(row=0, column=col, padx=4, sticky="ew")
            score_frame.columnconfigure(col, weight=1)
            ttk.Label(cell, text=title).pack(anchor="w")
            var = tk.StringVar(value="")
            self.score_vars[key] = var
            entry = ttk.Entry(
                cell,
                textvariable=var,
                width=10,
                validate="key",
                validatecommand=vcmd,
            )
            entry.pack(fill=tk.X)
            # 禁用滚轮改分（避免误触）
            entry.bind("<MouseWheel>", lambda e: "break")
            entry.bind("<Button-4>", lambda e: "break")
            entry.bind("<Button-5>", lambda e: "break")
            self.score_widgets[key] = entry

    def pack(self, **kwargs) -> None:
        self.frame.pack(**kwargs)

    @staticmethod
    def _validate_score_key(new_value: str) -> bool:
        """输入过程允许空、中间态；完整值再严格校验。"""
        if new_value == "" or new_value in (".", "-", "0.", "1.", "2.", "3.", "4.", "5."):
            return True
        try:
            v = float(new_value)
        except ValueError:
            return False
        if v < SCORE_MIN or v > SCORE_MAX + 1e-9:
            return False
        # 最多一位小数
        if "." in new_value:
            frac = new_value.split(".", 1)[1]
            if len(frac) > 1:
                return False
        return True

    def _progress_pos_from_event(self, event) -> float:
        """按点击/拖动的 x 坐标换算为进度秒数（不换控件）。"""
        width = max(int(self.progress.winfo_width()), 1)
        # 两端略留白，减少点到滑块边缘时的偏差
        pad = 8
        usable = max(width - 2 * pad, 1)
        frac = (event.x - pad) / usable
        frac = min(1.0, max(0.0, frac))
        vmin = float(self.progress.cget("from"))
        vmax = float(self.progress.cget("to"))
        if vmax < vmin:
            vmin, vmax = vmax, vmin
        return vmin + (vmax - vmin) * frac

    def _apply_progress_pos(self, pos: float, *, seek_player: bool) -> None:
        pos = max(0.0, min(float(self.progress.cget("to")), float(pos)))
        self.progress.set(pos)
        self._refresh_time(pos)
        if seek_player and self._active:
            self.player.seek(pos)

    def _seek_press(self, event) -> None:
        self._seeking = True
        pos = self._progress_pos_from_event(event)
        # 按下时先跳到点击位置（点槽即可跳转）
        self._apply_progress_pos(pos, seek_player=False)

    def _seek_motion(self, event) -> None:
        if not self._seeking:
            return
        pos = self._progress_pos_from_event(event)
        self._apply_progress_pos(pos, seek_player=False)

    def _seek_release(self, event=None) -> None:
        if not self._seeking:
            return
        self._seeking = False
        if event is not None:
            pos = self._progress_pos_from_event(event)
        else:
            pos = float(self.progress.get())
        # 松开时同步播放器位置：播放中会跳转；暂停中只改位置不播
        self._apply_progress_pos(pos, seek_player=True)

    def _on_seek_drag(self, value: str) -> None:
        # 保留 Scale 自身拖滑块时的回调；与坐标换算互补
        if not self._seeking:
            return
        try:
            pos = float(value)
        except ValueError:
            return
        self._refresh_time(pos)

    def toggle_play(self) -> None:
        if self._active and self.player.is_playing():
            self.player.pause()
            self.btn_play.configure(text="▶ 继续")
            return
        if self._active and self.player.is_paused():
            self.player.unpause()
            self.btn_play.configure(text="⏸ 暂停")
            return
        # 请求成为当前播放行
        self.on_play_request(self)

    def start_playback(self, from_pos: Optional[float] = None) -> None:
        pos = float(self.progress.get()) if from_pos is None else from_pos
        # 已播到末尾时再次播放，从 0 重播
        if self.duration > 0 and pos >= max(0.0, self.duration - 0.05):
            pos = 0.0
            self.progress.set(0.0)
            self._refresh_time(0.0)
        try:
            self.player.load(self.path, self.duration)
            self.player.play(pos)
        except Exception as exc:
            self._active = False
            self.btn_play.configure(text="▶ 播放")
            messagebox.showerror("播放失败", f"{self.path.name}\n{exc}")
            return
        self._active = True
        self.btn_play.configure(text="⏸ 暂停")

    def deactivate(self) -> None:
        self._active = False
        self.btn_play.configure(text="▶ 播放")
        self.progress.set(0.0)
        self._refresh_time(0.0)

    def on_stopped_externally(self) -> None:
        self._active = False
        self.btn_play.configure(text="▶ 播放")

    def tick(self) -> None:
        if not self._active or self._seeking:
            return
        if self.player.is_paused():
            return
        pos = self.player.get_position()
        if not self.player.is_playing():
            # 播放结束
            pos = self.duration
            self.progress.set(pos)
            self._refresh_time(pos)
            self.btn_play.configure(text="▶ 播放")
            self._active = False
            return
        self.progress.set(pos)
        self._refresh_time(pos)

    def _refresh_time(self, pos: float) -> None:
        self.time_var.set(f"{format_time(pos)} / {format_time(self.duration)}")

    def get_scores(self) -> Optional[Dict[str, float]]:
        scores: Dict[str, float] = {}
        for key, _title in self.dimensions:
            text = self.score_vars[key].get().strip()
            if not is_valid_score(text):
                return None
            scores[key] = quantize_score(float(text))
        return scores

    def missing_or_invalid_dims(self) -> List[str]:
        bad: List[str] = []
        for key, title in self.dimensions:
            text = self.score_vars[key].get().strip()
            if not is_valid_score(text):
                bad.append(title)
        return bad


# =============================================================================
# 评分标准侧栏面板（常驻显示）
# =============================================================================

class ScoreRubricPanel(ttk.Frame):
    """在打分界面右侧常驻展示当前场景评分标准。"""

    def __init__(self, master: tk.Widget, scene_index: int, **kwargs) -> None:
        super().__init__(master, **kwargs)
        rubric = rubric_for_scene(scene_index)
        title = str(rubric.get("title", "评分标准"))
        scene_name = (
            SCENE_DISPLAY_NAMES[scene_index]
            if scene_index < len(SCENE_DISPLAY_NAMES)
            else f"场景{scene_index + 1}"
        )
        headers = list(rubric.get("headers") or [])
        rows = list(rubric.get("rows") or [])
        col_count = max(len(headers), 1)
        body_labels: List[ttk.Label] = []

        ttk.Label(self, text=title, font=("", 12, "bold")).pack(anchor="w", pady=(0, 2))
        ttk.Label(
            self,
            text=f"{scene_name}\n分值 {SCORE_MIN}–{SCORE_MAX}，步进 {SCORE_STEP}",
            foreground="#444",
            justify=tk.LEFT,
        ).pack(anchor="w", pady=(0, 8))

        table_wrap = ttk.Frame(self)
        table_wrap.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(table_wrap, highlightthickness=0)
        vsb = ttk.Scrollbar(table_wrap, orient=tk.VERTICAL, command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas_window = canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)

        def _on_canvas_configure(event):
            canvas.itemconfigure(canvas_window, width=event.width)
            new_wrap = max(120, (event.width - 100) // max(col_count - 1, 1))
            for lbl in body_labels:
                lbl.configure(wraplength=new_wrap)

        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        # 右侧评分标准表禁用滚轮滚动，仅使用滚动条拖动

        for c, text in enumerate(headers):
            cell = ttk.Label(
                inner,
                text=str(text),
                font=("", 9, "bold"),
                borderwidth=1,
                relief="solid",
                padding=4,
                anchor="center",
            )
            cell.grid(row=0, column=c, sticky="nsew")
            inner.columnconfigure(c, weight=1 if c > 0 else 0, minsize=72 if c == 0 else 120)

        wrap = 160 if col_count >= 3 else 220
        for r, row in enumerate(rows, start=1):
            row_vals = list(row) if isinstance(row, (list, tuple)) else [row]
            while len(row_vals) < col_count:
                row_vals.append("")
            for c in range(col_count):
                text = str(row_vals[c])
                cell = ttk.Label(
                    inner,
                    text=text,
                    borderwidth=1,
                    relief="solid",
                    padding=4,
                    anchor="nw" if c > 0 else "center",
                    wraplength=wrap if c > 0 else 72,
                    justify=tk.LEFT if c > 0 else tk.CENTER,
                )
                cell.grid(row=r, column=c, sticky="nsew")
                if c > 0:
                    body_labels.append(cell)


# =============================================================================
# 主应用
# =============================================================================

class BlindListenApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("盲听音频评分工具")
        self.geometry("1500x800")
        self.minsize(1280, 700)

        self.evaluator_name = ""
        self.scene_index = 0
        self.player: Optional[WavPlayer] = None
        self.rows: List[AudioRow] = []
        self.active_row: Optional[AudioRow] = None
        # results: list of dict records
        self.results: List[dict] = []

        try:
            self.player = WavPlayer()
        except (AudioDeviceError, pygame.error) as exc:
            show_audio_device_error(exc)
            self.destroy()
            raise SystemExit(1) from exc

        self._build_style()
        self._show_name_screen()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.after(PROGRESS_TICK_MS, self._progress_loop)

    def _build_style(self) -> None:
        style = ttk.Style(self)
        # 尽量使用系统原生主题
        for theme in ("vista", "aqua", "clam", style.theme_use()):
            if theme in style.theme_names():
                try:
                    style.theme_use(theme)
                    break
                except tk.TclError:
                    continue

    # ----- 姓名页 -----
    def _show_name_screen(self) -> None:
        self._clear_root()
        frame = ttk.Frame(self, padding=40)
        frame.pack(expand=True)

        ttk.Label(frame, text="盲听音频评分", font=("", 18, "bold")).pack(pady=(0, 12))
        ttk.Label(
            frame,
            text="开始前请输入评分人姓名。测评过程中界面仅显示音频编号，不显示真实文件名。测评完成后，将在当前目录下生成一个Excel文件。",
            wraplength=520,
            justify=tk.CENTER,
        ).pack(pady=(0, 24))

        form = ttk.Frame(frame)
        form.pack()
        ttk.Label(form, text="姓名：").grid(row=0, column=0, sticky="e", padx=(0, 8))
        self.name_var = tk.StringVar()
        entry = ttk.Entry(form, textvariable=self.name_var, width=28)
        entry.grid(row=0, column=1)
        entry.focus_set()
        entry.bind("<Return>", lambda _e: self._start_eval())

        ttk.Button(frame, text="开始测评", command=self._start_eval).pack(pady=24)

        # tip = (
        #     # "音频已内置于程序中（打包后不可从外部替换）。\n"
        #     f"场景: {', '.join(SCENE_FOLDERS)}（各约 {NUM_AUDIO_PER_SCENE} 条 wav）"
        # )
        tip = (
            f"一共8种场景（各 {NUM_AUDIO_PER_SCENE} 条 wav）, 预计需要30分钟完成"
        )
        # if not getattr(sys, "frozen", False):
        #     tip = (
        #         # f"音频目录: {music_root()}\n"
        #         f"场景: {', '.join(SCENE_FOLDERS)}（各约 {NUM_AUDIO_PER_SCENE} 条 wav）"
        #     )
        if DEBUG_COPY_FIRST_SCENE_SCORES:
            tip += (
                "\n\n【调试模式已开启】填完第1个场景后，"
                "后续场景评分将按同编号自动复制并导出 Excel。"
                "\n正式测评请将 DEBUG_COPY_FIRST_SCENE_SCORES 改为 0。"
            )
        ttk.Label(frame, text=tip, foreground="#555", justify=tk.CENTER).pack()

    def _start_eval(self) -> None:
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("提示", "请输入评分人姓名。")
            return
        self.evaluator_name = name

        # 启动前检查目录
        missing = []
        for folder in SCENE_FOLDERS:
            wavs = list_scene_wavs(folder)
            if len(wavs) == 0:
                missing.append(f"{folder}/ （无 wav）")
            elif len(wavs) != NUM_AUDIO_PER_SCENE:
                missing.append(
                    f"{folder}/ （找到 {len(wavs)} 个，期望 {NUM_AUDIO_PER_SCENE} 个）"
                )
        if any("无 wav" in m for m in missing):
            messagebox.showerror(
                "音频目录不完整",
                "以下场景缺少音频，请检查后重试：\n\n" + "\n".join(missing),
            )
            return
        if missing:
            ok = messagebox.askyesno(
                "音频数量提示",
                "以下场景音频数量与期望不符，是否仍继续？\n\n" + "\n".join(missing),
            )
            if not ok:
                return

        self.scene_index = 0
        self.results.clear()
        self._show_scene_screen()

    # ----- 场景页 -----
    def _show_scene_screen(self) -> None:
        self._stop_playback()
        self._clear_root()

        scene_folder = SCENE_FOLDERS[self.scene_index]
        display = SCENE_DISPLAY_NAMES[self.scene_index] if self.scene_index < len(
            SCENE_DISPLAY_NAMES
        ) else scene_folder
        wavs = list_scene_wavs(scene_folder)

        header = ttk.Frame(self, padding=(12, 10))
        header.pack(fill=tk.X)
        ttk.Label(
            header,
            text=f"评分人：{self.evaluator_name}",
        ).pack(side=tk.LEFT)
        ttk.Label(
            header,
            text=f"{display}  "
                 f"{self.scene_index + 1}/{len(SCENE_FOLDERS)}",
            font=("", 12, "bold"),
        ).pack(side=tk.LEFT, padx=24)
        dims = dimensions_for_scene(self.scene_index)
        dim_names = "、".join(title for _k, title in dims)
        ttk.Label(
            header,
            text="请为每条音频打分：0.0–5.0，步进 0.1（请手动输入）",
        ).pack(side=tk.RIGHT)

        # 左右分栏：左侧音频打分，右侧评分标准（常驻）
        paned = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

        left = ttk.Frame(paned)
        right_box = ttk.LabelFrame(paned, text="评分标准（常驻参考）", padding=8)
        paned.add(left, weight=3)
        paned.add(right_box, weight=2)

        ScoreRubricPanel(right_box, self.scene_index).pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(left, highlightthickness=0)
        scrollbar = ttk.Scrollbar(left, orient=tk.VERTICAL, command=canvas.yview)
        self.scroll_inner = ttk.Frame(canvas)

        self.scroll_inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        self._canvas_window = canvas.create_window((0, 0), window=self.scroll_inner, anchor="nw")
        canvas.bind(
            "<Configure>",
            lambda e: canvas.itemconfigure(self._canvas_window, width=e.width),
        )
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 左侧音频区滚轮：仅鼠标在左侧时生效，避免右侧也被带动
        def _on_mousewheel(event):
            if sys.platform == "darwin":
                canvas.yview_scroll(-1 * int(event.delta), "units")
            else:
                canvas.yview_scroll(-1 * int(event.delta / 120), "units")

        def _on_linux_scroll(event):
            canvas.yview_scroll(-1 if event.num == 4 else 1, "units")

        def _bind_left_wheel(_event=None):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_linux_scroll)
            canvas.bind_all("<Button-5>", _on_linux_scroll)

        def _unbind_left_wheel(_event=None):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        left.bind("<Enter>", _bind_left_wheel)
        left.bind("<Leave>", _unbind_left_wheel)
        # 进入右侧时明确解绑，防止左侧 bind_all 仍滚动左侧列表
        right_box.bind("<Enter>", _unbind_left_wheel)
        self._canvas = canvas
        _bind_left_wheel()

        self.rows = []
        for i, path in enumerate(wavs):
            row = AudioRow(
                self.scroll_inner,
                i,
                path,
                self.player,
                self._on_play_request,
                dimensions=dims,
            )
            row.pack(fill=tk.X, padx=4, pady=6)
            self.rows.append(row)

        # 初始分割比例：左侧约 58%，右侧约 42%
        def _set_sash():
            try:
                total = paned.winfo_width()
                if total > 100:
                    paned.sashpos(0, int(total * 0.58))
            except tk.TclError:
                pass

        self.after(50, _set_sash)

        footer = ttk.Frame(self, padding=12)
        footer.pack(fill=tk.X)
        ttk.Button(footer, text="退出", command=self._on_close).pack(side=tk.LEFT)

        is_last = self.scene_index >= len(SCENE_FOLDERS) - 1
        if DEBUG_COPY_FIRST_SCENE_SCORES and self.scene_index == 0:
            next_text = "完成并导出 Excel（调试：复制到其余场景）"
        elif is_last:
            next_text = "完成并导出 Excel"
        else:
            next_text = "下一场景"
        ttk.Button(footer, text=next_text, command=self._next_scene).pack(side=tk.RIGHT)

    def _on_play_request(self, row: AudioRow) -> None:
        if self.active_row is not None and self.active_row is not row:
            self.player.stop()
            self.active_row.on_stopped_externally()
            self.active_row.progress.set(0.0)
            self.active_row._refresh_time(0.0)
        self.active_row = row
        row.start_playback()

    def _stop_playback(self) -> None:
        self.player.stop()
        if self.active_row is not None:
            self.active_row.on_stopped_externally()
            self.active_row = None

    def _progress_loop(self) -> None:
        if self.active_row is not None:
            self.active_row.tick()
            if not self.active_row._active:
                self.active_row = None
        self.after(PROGRESS_TICK_MS, self._progress_loop)

    def _collect_current_scene(self) -> bool:
        scene_folder = SCENE_FOLDERS[self.scene_index]
        display = SCENE_DISPLAY_NAMES[self.scene_index] if self.scene_index < len(
            SCENE_DISPLAY_NAMES
        ) else scene_folder

        for row in self.rows:
            bad = row.missing_or_invalid_dims()
            if bad:
                messagebox.showwarning(
                    "评分不完整或非法",
                    f"音频 {row.code} 以下维度无效（需 0.0–5.0，步进 0.1）：\n"
                    + "、".join(bad),
                )
                return False

        for row in self.rows:
            scores = row.get_scores()
            assert scores is not None
            record = {
                "evaluator": self.evaluator_name,
                "scene_folder": scene_folder,
                "scene_name": display,
                "scene_type": scene_type_label(self.scene_index),
                "code": row.code,
                "filename": row.path.name,
                "relpath": audio_relpath(row.path),
            }
            # 全部评分列先置空，再写入本场景实际维度
            for key, _title in ALL_SCORE_DIMENSIONS:
                record[key] = ""
            record.update(scores)
            self.results.append(record)
        return True

    def _scores_by_code_from_first_scene(self) -> Dict[str, Dict[str, float]]:
        """从已收集的第1个场景结果中，按代号取出已填分数。"""
        first_folder = SCENE_FOLDERS[0]
        mapping: Dict[str, Dict[str, float]] = {}
        for rec in self.results:
            if rec["scene_folder"] != first_folder:
                continue
            vals: Dict[str, float] = {}
            for key, _title in ALL_SCORE_DIMENSIONS:
                raw = rec.get(key, "")
                if raw == "" or raw is None:
                    continue
                vals[key] = float(raw)
            mapping[rec["code"]] = vals
        return mapping

    def _auto_fill_remaining_scenes_from_first(self) -> None:
        """调试：把第1个场景的同编号评分复制到后续全部场景。"""
        score_map = self._scores_by_code_from_first_scene()
        if not score_map:
            raise RuntimeError("调试复制失败：第1个场景没有可用评分。")

        for idx in range(1, len(SCENE_FOLDERS)):
            scene_folder = SCENE_FOLDERS[idx]
            display = (
                SCENE_DISPLAY_NAMES[idx]
                if idx < len(SCENE_DISPLAY_NAMES)
                else scene_folder
            )
            target_dims = dimensions_for_scene(idx)
            wavs = list_scene_wavs(scene_folder)
            for i, path in enumerate(wavs):
                code = code_label(i)
                src_scores = score_map.get(code)
                if src_scores is None:
                    codes = list(score_map.keys())
                    src_scores = score_map[codes[i % len(codes)]]
                # 维度不一致时（人声→噪声），用源场景已有分数填入目标各维，便于测表
                seed = next(iter(src_scores.values()), 3.0)
                record = {
                    "evaluator": self.evaluator_name,
                    "scene_folder": scene_folder,
                    "scene_name": display,
                    "scene_type": scene_type_label(idx),
                    "code": code,
                    "filename": path.name,
                    "relpath": audio_relpath(path),
                }
                for key, _title in ALL_SCORE_DIMENSIONS:
                    record[key] = ""
                for key, _title in target_dims:
                    record[key] = src_scores[key] if key in src_scores else seed
                self.results.append(record)

    def _next_scene(self) -> None:
        if not self._collect_current_scene():
            return
        self._stop_playback()

        # 调试宏：第1场景填完后，自动复制并导出
        if DEBUG_COPY_FIRST_SCENE_SCORES and self.scene_index == 0:
            try:
                self._auto_fill_remaining_scenes_from_first()
            except Exception as exc:
                messagebox.showerror("调试复制失败", str(exc))
                return
            messagebox.showinfo(
                "调试模式",
                "已将第1个场景的评分按同编号复制到其余场景，即将导出 Excel。\n"
                "正式测评请把 DEBUG_COPY_FIRST_SCENE_SCORES 改为 0。",
            )
            self._export_excel()
            return

        if self.scene_index >= len(SCENE_FOLDERS) - 1:
            self._export_excel()
            return
        self.scene_index += 1
        self._show_scene_screen()

    def _export_excel(self) -> None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c for c in self.evaluator_name if c not in r'\/:*?"<>|').strip()
        out_name = f"{EXCEL_FILENAME_PREFIX}_{safe_name}_{ts}.xlsx"
        out_path = writable_dir() / out_name

        wb = Workbook()
        ws = wb.active
        ws.title = "评分明细"

        headers = [
            "评分人",
            "场景名称",
            "场景类型",
            "场景文件夹",
            "代号",
            "音频文件名",
            "相对路径",
        ] + [title for _k, title in ALL_SCORE_DIMENSIONS]

        ws.append(headers)
        for rec in self.results:
            row = [
                rec["evaluator"],
                rec["scene_name"],
                rec.get("scene_type", ""),
                rec["scene_folder"],
                rec["code"],
                rec["filename"],
                rec["relpath"],
            ]
            for key, _title in ALL_SCORE_DIMENSIONS:
                val = rec.get(key, "")
                row.append("" if val is None else val)
            ws.append(row)

        # 映射表：代号 <-> 真实文件名
        ws2 = wb.create_sheet("文件名映射")
        ws2.append(["场景名称", "场景类型", "场景文件夹", "代号", "音频文件名", "相对路径"])
        seen = set()
        for rec in self.results:
            key = (rec["scene_folder"], rec["code"])
            if key in seen:
                continue
            seen.add(key)
            ws2.append([
                rec["scene_name"],
                rec.get("scene_type", ""),
                rec["scene_folder"],
                rec["code"],
                rec["filename"],
                rec["relpath"],
            ])

        # 维度说明
        ws3 = wb.create_sheet("评分维度说明")
        ws3.append(["场景类型", "适用场景", "需评分维度", "分值范围"])
        ws3.append([
            "人声质量",
            SCENE_DISPLAY_NAMES[VOICE_QUALITY_SCENE_INDEX]
            if VOICE_QUALITY_SCENE_INDEX < len(SCENE_DISPLAY_NAMES)
            else "场景1",
            "、".join(t for _k, t in VOICE_SCENE_DIMENSIONS),
            f"{SCORE_MIN}–{SCORE_MAX}，步进 {SCORE_STEP}",
        ])
        noise_names = [
            SCENE_DISPLAY_NAMES[i]
            for i in range(len(SCENE_FOLDERS))
            if i != VOICE_QUALITY_SCENE_INDEX and i < len(SCENE_DISPLAY_NAMES)
        ]
        ws3.append([
            "噪声场景",
            "；".join(noise_names),
            "、".join(t for _k, t in NOISE_SCENE_DIMENSIONS),
            f"{SCORE_MIN}–{SCORE_MAX}，步进 {SCORE_STEP}",
        ])
        ws3.append([])
        ws3.append([
            "说明",
            "评分明细表中，本场景不适用的维度单元格为空。",
        ])

        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for sheet in (ws, ws2, ws3):
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
                for cell in row:
                    cell.border = thin
            for col in sheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    max_len = max(max_len, len(str(cell.value or "")))
                sheet.column_dimensions[col_letter].width = min(max_len + 4, 48)

        try:
            wb.save(str(out_path))
        except OSError as exc:
            messagebox.showerror("保存失败", str(exc))
            return

        messagebox.showinfo(
            "导出完成",
            f"共 {len(self.results)} 条评分已保存：\n{out_path}",
        )
        self._show_done_screen(out_path)

    def _show_done_screen(self, out_path: Path) -> None:
        self._stop_playback()
        self._clear_root()
        frame = ttk.Frame(self, padding=40)
        frame.pack(expand=True)
        ttk.Label(frame, text="测评完成", font=("", 18, "bold")).pack(pady=(0, 12))
        ttk.Label(
            frame,
            text=f"结果文件：\n{out_path}",
            justify=tk.CENTER,
        ).pack(pady=(0, 20))
        ttk.Button(frame, text="再评一次", command=self._show_name_screen).pack(pady=4)
        ttk.Button(frame, text="退出", command=self._on_close).pack(pady=4)

    def _clear_root(self) -> None:
        try:
            if hasattr(self, "_canvas"):
                self.unbind_all("<MouseWheel>")
                self.unbind_all("<Button-4>")
                self.unbind_all("<Button-5>")
        except tk.TclError:
            pass
        for child in self.winfo_children():
            child.destroy()
        self.rows = []
        self.active_row = None

    def _on_close(self) -> None:
        if self.results and self.scene_index < len(SCENE_FOLDERS) - 1:
            # 可能有未完成数据；已写入 results 的场景已收集
            pass
        if messagebox.askokcancel("退出", "确定退出？未完成导出的评分将丢失。"):
            if self.player is not None:
                self.player.shutdown()
            self.destroy()


def ensure_music_dirs() -> None:
    """仅开发模式下创建 music 目录骨架；打包后音频只读内嵌，不在 exe 旁生成可改目录。"""
    if getattr(sys, "frozen", False):
        return
    root = music_root()
    root.mkdir(parents=True, exist_ok=True)
    for folder in SCENE_FOLDERS:
        (root / folder).mkdir(parents=True, exist_ok=True)


def probe_audio_device() -> None:
    """启动前探测音频设备；失败则弹窗并退出进程。"""
    try:
        init_pygame_mixer()
        try:
            pygame.mixer.quit()
        except Exception:
            pass
    except (AudioDeviceError, pygame.error, OSError, TimeoutError) as exc:
        show_audio_device_error(exc)
        sys.exit(1)


def main() -> None:
    ensure_music_dirs()
    # 先探测音频，避免无设备时主界面卡死
    probe_audio_device()
    try:
        app = BlindListenApp()
    except SystemExit:
        raise
    except (AudioDeviceError, pygame.error) as exc:
        show_audio_device_error(exc)
        sys.exit(1)
    except Exception as exc:
        show_audio_device_error(exc)
        sys.exit(1)
    app.mainloop()


if __name__ == "__main__":
    main()
